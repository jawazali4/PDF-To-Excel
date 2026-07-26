"""
Excel writer with full Arabic support.
Fixes Arabic text shape/direction before writing.
All values written as TEXT to prevent any data changes.
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from utils.config import EXCEL_FONT_NAME, EXCEL_FONT_SIZE, MAX_COLUMN_WIDTH
from utils.arabic_utils import fix_arabic_cell, fix_arabic_table, fix_arabic_dict, _contains_arabic


class ExcelWriter:
    """Write data to Excel with Arabic support and data preservation."""

    def __init__(self):
        self.wb = Workbook()
        self.wb.remove(self.wb.active)
        self._sheet_counter = 0

        # Styles
        self._header_fill  = PatternFill(start_color="1a3a6b", end_color="1a3a6b", fill_type="solid")
        self._header_font  = Font(name=EXCEL_FONT_NAME, size=EXCEL_FONT_SIZE, bold=True, color="FFFFFF")
        self._cell_font    = Font(name=EXCEL_FONT_NAME, size=EXCEL_FONT_SIZE)
        self._rtl_align    = Alignment(horizontal="right", vertical="top", wrap_text=True)
        self._ltr_align    = Alignment(horizontal="left",  vertical="top", wrap_text=True)
        self._center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self._thin_border  = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"),  bottom=Side(style="thin"),
        )

    # ──────────────────────────────────────────────────────
    def add_table_sheet(self, name: str, table_data: list, has_header: bool = True):
        """Add a sheet with table data — Arabic text fixed automatically."""
        self._sheet_counter += 1
        ws = self._create_sheet(name)

        # Fix Arabic in the whole table at once
        fixed_table = fix_arabic_table(table_data)

        for r_idx, row in enumerate(fixed_table, start=1):
            for c_idx, value in enumerate(row if row else [], start=1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value         = self._safe_str(value)
                cell.number_format = "@"           # Force text — no auto-convert
                cell.border        = self._thin_border

                if has_header and r_idx == 1:
                    cell.font      = self._header_font
                    cell.fill      = self._header_fill
                    cell.alignment = self._center_align
                else:
                    cell.font      = self._cell_font
                    # Align based on content direction
                    cell.alignment = (
                        self._rtl_align
                        if _contains_arabic(str(value))
                        else self._ltr_align
                    )

        self._autofit(ws)
        if has_header and fixed_table:
            ws.freeze_panes = "A2"

    # ──────────────────────────────────────────────────────
    def add_text_sheet(self, name: str, lines: list):
        """Add a sheet with text lines — Arabic fixed."""
        self._sheet_counter += 1
        ws = self._create_sheet(name)

        for r_idx, line in enumerate(lines, start=1):
            fixed = fix_arabic_cell(str(line)) if line else ""
            cell  = ws.cell(row=r_idx, column=1)
            cell.value         = fixed
            cell.number_format = "@"
            cell.font          = self._cell_font
            cell.alignment     = (
                self._rtl_align
                if _contains_arabic(fixed)
                else self._ltr_align
            )

        self._autofit(ws)

    # ──────────────────────────────────────────────────────
    def add_key_value_sheet(self, name: str, data: dict):
        """Add key-value sheet (invoice fields) — Arabic fixed."""
        self._sheet_counter += 1
        ws = self._create_sheet(name)

        # Header row
        for c_idx, header in enumerate(["Field", "Value"], start=1):
            cell           = ws.cell(row=1, column=c_idx, value=header)
            cell.font      = self._header_font
            cell.fill      = self._header_fill
            cell.alignment = self._center_align
            cell.border    = self._thin_border

        # Fix Arabic in dict
        fixed_data = fix_arabic_dict(data)

        row = 2
        for key, value in fixed_data.items():
            # Key cell
            key_cell               = ws.cell(row=row, column=1, value=str(key))
            key_cell.number_format = "@"
            key_cell.font          = Font(name=EXCEL_FONT_NAME, size=EXCEL_FONT_SIZE, bold=True)
            key_cell.alignment     = self._rtl_align if _contains_arabic(str(key)) else self._ltr_align
            key_cell.border        = self._thin_border

            # Value cell
            val_str                = str(value) if value else ""
            val_cell               = ws.cell(row=row, column=2, value=val_str)
            val_cell.number_format = "@"
            val_cell.font          = self._cell_font
            val_cell.alignment     = self._rtl_align if _contains_arabic(val_str) else self._ltr_align
            val_cell.border        = self._thin_border

            row += 1

        self._autofit(ws)
        ws.freeze_panes = "A2"

    # ──────────────────────────────────────────────────────
    def add_multi_tables_sheet(self, name: str, tables: list):
        """Multiple tables in one sheet — Arabic fixed."""
        self._sheet_counter += 1
        ws = self._create_sheet(name)
        current_row = 1

        for t_idx, table in enumerate(tables):
            # Table label
            label      = ws.cell(row=current_row, column=1, value=f"Table {t_idx + 1}")
            label.font = Font(name=EXCEL_FONT_NAME, size=12, bold=True, color="1a3a6b")
            current_row += 1

            fixed_table = fix_arabic_table(table)

            for row in fixed_table:
                for c_idx, value in enumerate(row if row else [], start=1):
                    cell               = ws.cell(row=current_row, column=c_idx)
                    cell.value         = self._safe_str(value)
                    cell.number_format = "@"
                    cell.font          = self._cell_font
                    cell.alignment     = (
                        self._rtl_align
                        if _contains_arabic(str(value))
                        else self._ltr_align
                    )
                    cell.border = self._thin_border
                current_row += 1

            current_row += 2  # spacing between tables

        self._autofit(ws)

    # ──────────────────────────────────────────────────────
    def save(self, path: str):
        """Save the workbook."""
        if not self.wb.sheetnames:
            ws          = self.wb.create_sheet(title="Empty")
            ws["A1"]    = "No data was extracted."
            ws["A1"].font = self._cell_font
        self.wb.save(path)

    # ══════════════════════════════════════════════════════
    #  Private Helpers
    # ══════════════════════════════════════════════════════
    def _create_sheet(self, name: str):
        """Create an RTL worksheet."""
        import re
        clean = re.sub(r'[:\\/*?\[\]]', "_", str(name))[:31]
        if clean in self.wb.sheetnames:
            clean = f"{clean[:27]}_{self._sheet_counter}"
        ws = self.wb.create_sheet(title=clean)
        ws.sheet_view.rightToLeft = True   # RTL for Arabic
        return ws

    @staticmethod
    def _safe_str(value) -> str:
        """Safe string conversion."""
        if value is None:
            return ""
        return str(value).replace("\u00A0", " ").replace("\ufeff", "").strip()

    @staticmethod
    def _autofit(ws):
        """Auto-fit column widths."""
        for col_cells in ws.columns:
            max_len    = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                val     = str(cell.value) if cell.value else ""
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = min(max_len + 4, MAX_COLUMN_WIDTH)